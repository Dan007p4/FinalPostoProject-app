import pandas as pd
import streamlit as st
from streamlit_option_menu import option_menu
import streamlit_authenticator as stauth
import yaml
import xlrd
import mysql.connector
from yaml.loader import SafeLoader
from sqlalchemy import create_engine
import seaborn as sns
import matplotlib.pyplot as plt
import unidecode
import openpyxl
import time
import datetime as dt
from io import BytesIO
import xlsxwriter
import extra_streamlit_components as stx
from pyxlsb import open_workbook as open_xlsb
from datetime import datetime, timedelta
from streamlit_cookies_manager import EncryptedCookieManager
import bcrypt
import math
import calendar
import time
from fpdf import FPDF
import base64


st.set_page_config(page_icon="🏥", page_title="Gerenciador de dados")
##FAZENDO CONEXÃO COM O DB##

connection = mysql.connector.connect(
    host="aws.connect.psdb.cloud",
    user=st.secrets["db_username"],
    passwd=st.secrets["db_password"],

    db="database",
    ssl_ca="cacert-2023-01-10.pem"


)


c = connection.cursor()


def get_manager():
    return stx.CookieManager()


cookie_manager = get_manager()
cookiee = "ActualUser"

if 'Login2' not in st.session_state:
    st.session_state['Login2'] = 0


if 'ActualUser' not in st.session_state:
    st.session_state['ActualUser'] = 0

# cookie_manager.set(cookiee, None, expires_at=datetime.now() + timedelta(days=30))


if((cookie_manager.get(cookie=cookiee)) == 'null'):
    st.session_state.Login2 = 0

if ((cookie_manager.get(cookie=cookiee)) != 'null'):
    if((str(cookie_manager.get(cookie=cookiee)).split("|")[0]) == "0"):  
        st.session_state.Login2 = 0

    if((str(cookie_manager.get(cookie=cookiee)).split("|")[0]) == "1"):
        st.session_state.Login2 = 1

    if((str(cookie_manager.get(cookie=cookiee)).split("|")[0]) == "2"):
        st.session_state.Login2 = 2
    if (len(str(cookie_manager.get( cookie=cookiee)).split("|")) == 1 ):
        st.session_state['ActualUser'] = str(cookie_manager.get(
            cookie=cookiee)).split("|")[0]
    else:
        st.session_state['ActualUser'] = str(cookie_manager.get(
            cookie=cookiee)).split("|")[1]
    

if(((st.session_state.Login2 == 0) | (st.session_state.Login2 == 3))):
    # st.write(cookie_manager.get(cookie= cookiee))
    # st.write(st.session_state.Login2)

    cookie = "ActualUser"
    date_current = datetime.now()
    st.title("Login")

    user = st.text_input("Usuário")
    password = st.text_input("Senha", type="password")
    css = '''
            <style>
            [class="css-1li7dat effi0qh1"]{visibility: hidden;}
            </style>
            '''
    st.markdown(css, unsafe_allow_html=True)

    butt = st.button("Login")

    c.execute(
        "SELECT NOME_LOGIN FROM TABELA_LOGINS_POSTO ;")

    logins_Name = c.fetchall()

    c.execute(
        "SELECT SENHA_LOGIN FROM TABELA_LOGINS_POSTO WHERE NOME_LOGIN ='" + str(user)+"';")

    logins_Pass = c.fetchall()

    salt = st.secrets["salt"]

    test = bcrypt.hashpw(password.encode('utf-8'), salt)
    new = bcrypt.hashpw(password.encode('utf-8'), salt)
    verify = False
    count = -1
    if butt:
        if len(logins_Pass) != 0:

            for i in logins_Name:
                if verify != True:
                    count = count+1
                    if((str(user) == str(i[0])) & (str(test).replace("'",":") == str(logins_Pass[0][0]))):
                        verify = True
                    else:
                        verify = False

            if(verify == True):
                c.execute(
                    "SELECT NIVEL_PERMISSAO FROM TABELA_LOGINS_POSTO WHERE NOME_LOGIN ='" + str(user)+"';")

                logins_Perm = c.fetchall()
                st.session_state.Login2 = int(logins_Perm[0][0])

                c.execute(
                    "SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO WHERE NOME_LOGIN ='" + str(user)+"';")

                logins_UserId = c.fetchall()
                st.session_state.ActualUser = int(logins_UserId[0][0])

                cookie_manager.set(cookie, str(
                    logins_Perm[0][0])+"|"+str(logins_UserId[0][0]), expires_at=datetime.now() + timedelta(days=5))
                time.sleep(1000)
                st.experimental_rerun()
            else:
                st.session_state.Login2 = 3
        else:
            st.error('Senha ou Usuario esta incorreto')


##CRIANDO MENU##

salt = b'$2b$12$9FEmZ.X34ET241I6wiYPze'
def Clean_Names(name):
    name = str(name)
    name = unidecode.unidecode(name)
    name = name.replace(" ", '_')
    name = name.replace("/", '_')
    name = name.replace(".", '')

    return name


def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'})
    worksheet.set_column('A:A', None, format1)
    writer.close()
    processed_data = output.getvalue()
    return processed_data


meses_dict = {
    "JANEIRO": "01",
    "FEVEREIRO": "02",
    "MARÇO": "03",
    "ABRIL": "04",
    "MAIO": "05",
    "JUNHO": "06",
    "JULHO": "07",
    "AGOSTO": "08",
    "SETEMBRO": "09",
    "OUTUBRO": "10",
    "NOVEMBRO": "11",
    "DEZEMBRO": "12"
}


def LogOut():
    st.session_state.Login2 = 0
    cookie_manager.set(
        cookiee, None, expires_at=datetime.now() + timedelta(days=30))
    time.sleep(1000)

    st.experimental_rerun()


def verificar_formato_data(data_string):
    caracteres_permitidos = ["0123456789-"]
    for caractere in data_string:
        if caractere not in caracteres_permitidos[0]:
            st.error("Data está contendo caracteres proibidos")
            return False
        if len(data_string) < 10:
            st.error("Data está incompleta")
            return False
    return True


def verificar_formato_CNS(data_string):
    caracteres_permitidos = ["0123456789"]
    for caractere in data_string:
        if caractere not in caracteres_permitidos[0]:
            st.error("CNS está contendo caracteres proibidos")
            return False
        if data_string =="":
            return False
        if data_string == None:
            return False
        if len(data_string) == 0:
            return False 
    return True


def verificar_formato_nome(data_string):
    symbols_and_accents = [
        "a", "á", "A", "Á", "à", "â", "ã", "ä", "b", "B",
        "c", "C", "d", "D", "e", "é", "è", "ê", "ë", "E",
        "f", "F", "g", "G", "h", "H", "i", "í", "ì", "î","I",
        "ï", "j", "J", "k", "K", "l", "L", "m", "M", "n",
        "N", "o", "ó", "ò", "ô", "õ", "ö", "O", "Ó", "Ò",
        "Ô", "Õ", "Ö", "p", "P", "q", "Q", "r", "R", "s",
        "S", "t", "T", "u", "ú", "ù", "û", "ü", "U", "Ú",
        "Ù", "Û", "Ü", "v", "V", "w", "W", "x", "X", "y",
        "ý", "Y", "Ý", "z", "Z", "_", "", " ", "ç", "Ç"
    ]

    for caractere in data_string:
        if caractere not in symbols_and_accents:
            st.error("Nome está contendo caracteres proibidos")
            return False
        if data_string =="":
            return False
        if data_string == None:
            return False
        if str(len(data_string)) == "0":
            return False 
    return True

def verificar_last_deleted(id):
    c.execute("SELECT ULTIMO_DELETADO FROM LAST_DELETED_POSTO ORDER BY ID DESC LIMIT 1")
    last_Deleted = c.fetchall()
    if len(last_Deleted)>0:
        if last_Deleted[0][0] == id:
            LogOut()
            

def verificar_formato_localOuEtio(data_string,value):
    symbols_and_accents = [
        "a", "á", "A", "Á", "à", "â", "ã", "ä", "b", "B",
        "c", "C", "d", "D", "e", "é", "è", "ê", "ë", "E",
        "f", "F", "g", "G", "h", "H", "i", "í", "ì", "î",
        "ï", "j", "J", "k", "K", "l", "L", "m", "M", "n",
        "N", "o", "ó", "ò", "ô", "õ", "ö", "O", "Ó", "Ò",
        "Ô", "Õ", "Ö", "p", "P", "q", "Q", "r", "R", "s",
        "S", "t", "T", "u", "ú", "ù", "û", "ü", "U", "Ú",
        "Ù", "Û", "Ü", "v", "V", "w", "W", "x", "X", "y",
        "ý", "Y", "Ý", "z", "Z", "_", "", " ", "ç", "Ç"
    ]

    for caractere in data_string:
        if caractere not in symbols_and_accents:
            st.error(value+" está contendo caracteres proibidos")
            return False
        if data_string =="":
            return False
        if data_string == None:
            return False
    if str(len(data_string)) == "0":
        return False 
    return True

def verificar_campo_texto(data_string):
    symbols_and_accents = [
        "a", "á", "A", "Á", "à", "â", "ã", "ä", "b", "B",
        "c", "C", "d", "D", "e", "é", "è", "ê", "ë", "E",
        "f", "F", "g", "G", "h", "H", "i", "í", "ì", "î","I",
        "ï", "j", "J", "k", "K", "l", "L", "m", "M", "n",
        "N", "o", "ó", "ò", "ô", "õ", "ö", "O", "Ó", "Ò",
        "Ô", "Õ", "Ö", "p", "P", "q", "Q", "r", "R", "s",
        "S", "t", "T", "u", "ú", "ù", "û", "ü", "U", "Ú",
        "Ù", "Û", "Ü", "v", "V", "w", "W", "x", "X", "y",
        "ý", "Y", "Ý", "z", "Z", "_", "", " ", "ç", "Ç"
    ]

    for caractere in data_string:
        if caractere not in symbols_and_accents:
            return False
        if data_string =="":
            return False
        if data_string == None:
            return False
        if str(len(data_string)) == "0":
            return False 
    return True

verificar_last_deleted(st.session_state.ActualUser)
# st.write(cookie_manager.get(
#         cookie=cookiee))
# st.write(st.session_state.ActualUser)
if(st.session_state.Login2 == 1):
    # authenticator.logout('Logout', 'main')
    with st.sidebar:

        selected = option_menu(
            menu_title="Menu",
            options=["Gerenciador de dados", "Analise de dados","Gerenciar Logins"],
            menu_icon="border-width"
        )

    st.sidebar.image(
        "WhatsApp Image 2023-02-21 at 14.22.25 (1).png", use_column_width=True)

    ##ABA GERENCIADOR DE DADOS##
    if selected == "Gerenciador de dados":
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=8)
        logout = st.button("Logout")
        if(logout):
            LogOut()

        # st.session_state.new_form2 = 0
        # st.divider()
        st.title("Gerenciador de dados")
        dateNow =datetime.now()
        
        c.execute(
            "SELECT UNIDADE_NOME FROM TABELA_LOGINS_POSTO;")
        list_tables = []
        list_months_verify = []
        list3_months_verify = []
        dict_total_material = {}
        list_years = ["2023"]
        
        tables = c.fetchall()

        for i in tables:
            value = i[0]
            list_tables.append(value)
        
        list_tables.append("Todas")

        unidade = st.selectbox("Selecione a Unidade",list_tables)
        filtro_relatorio = st.multiselect("Selecione os filtros do relatorio",['Paciente','Local da lesão','Codigo arqlife','Material','Quantidade'])
        filtro_mes = st.selectbox("Selecione o mês que deseja filtrar os pedidos",["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro","Todos"])
        for i in range(0,(dateNow.year - 2023)):
            list_years.append(str(2023+1))
        filtro_ano = st.selectbox("Selecione o ano que deseja filtrar os pedidos",list_years)
        c.execute(
            "SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO WHERE UNIDADE_NOME = '"+ str(unidade) +"' ;")

        if unidade == 'Todas':
            c.fetchall()
            c.execute(
            "SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO ;")
             
        list_tables = []
        tables = c.fetchall()
        for i in tables:
            value = i[0]
            list_tables.append(value)

        if filtro_mes =="Todos":

            for k in list_tables:
                numP = -1
                c.execute(
                "SELECT UNIDADE_NOME FROM TABELA_LOGINS_POSTO WHERE ID_UNIDADE = "+str(k)+" ;")
                nome= c.fetchall()
                st.divider()
                
                
                c.execute(
                    "SELECT * FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = "+ str(k) +" AND STATUS = 'ATIVO';")

                pacientes_unidades = c.fetchall()
                
                title   = "**Pacientes da unidade "+str(nome[0][0])+":**\n "


                for i in pacientes_unidades:
                    Nome = i[1]
                    Idade = i[3]
                    Genero = i[4]
                    Cns = i[5]
                    Status = i[7]
                    
                    title = title + "\n**● Nome do paciente: "+Nome+" | Idade paciente: "+str(Idade)+" | Gênero paciente : "+Genero+" | Cns do paciente : "+ str(Cns)+"**\n "
                st.subheader(title)
                st.divider()
                c.execute(
                    "SELECT * FROM TABELA_PROCESSOS_POSTO WHERE ID_UNIDADE = "+ str(k) +" AND DATA_DE_ENCERRAMENTO is null;")

                processos_unidades = c.fetchall()
                meses = {}
                meses["Janeiro"] = 1
                meses["Fevereiro"] = 2
                meses["Março"] = 3
                meses["Abril"] = 4
                meses["Maio"] = 5
                meses["Junho"] = 6
                meses["Julho"] = 7
                meses["Agosto"] = 8
                meses["Setembro"] = 9
                meses["Outubro"] = 10
                meses["Novembro"] = 11
                meses["Dezembro"] = 12


                title   = "**Processos da unidade "+str(nome[0][0])+":**\n "
                titleUtil = title.replace("*","")
                pdf.cell(200, 10, txt=titleUtil.replace("Processos","Pedidos"), ln=1, align="C")
                pdf.ln(2)
                st.subheader(title)
                for i in range(0,12):
                    numP = numP +1
                    p =list(meses)[numP]
                    st.divider()
                    st.subheader(p)  
                    for i in processos_unidades:
                        id_processo = i[0]
                        local = i[1]
                        etiologia = i[2]
                        dataA = i[5]
                        DataE = i[6]
                        Paciente = i[8]
                        if dateNow > ((dataA + timedelta(days=120))):
                            list3_months_verify.append(data.month)
                        c.execute(
                        "SELECT NOME,CNS FROM TABELA_PACIENTES_POSTO WHERE ID_PACIENTE = "+ str(Paciente) +" ;")

                        paciente_nome= c.fetchall()

                        # Criar um dicionário vazio
                        meses = {}

                        # Adicionar os meses e os números como pares de chave-valor
                        meses["Janeiro"] = 1
                        meses["Fevereiro"] = 2
                        meses["Março"] = 3
                        meses["Abril"] = 4
                        meses["Maio"] = 5
                        meses["Junho"] = 6
                        meses["Julho"] = 7
                        meses["Agosto"] = 8
                        meses["Setembro"] = 9
                        meses["Outubro"] = 10
                        meses["Novembro"] = 11
                        meses["Dezembro"] = 12

                        if filtro_mes =="Todos":
                                p =list(meses)[numP]
                                c.execute(
                                "SELECT * FROM TABELA_PEDIDOS_COBERTURAS_POSTO WHERE ID_PROCESSO = "+ str(id_processo) +" AND MONTH(DATA_ATUAL) ="+str(meses[p])+" AND YEAR(DATA_ATUAL) ="+str(filtro_ano)+" ;")

                                pedidios_processos = c.fetchall()
                                title2   = "**-Pedidos do processo:**\n "
                                for i in pedidios_processos:
                                    material = i[1]
                                    qnt = i[3]
                                    data = i[6]
                                    c.execute(
                                    "SELECT MODELO_NOME,COD_ARQ_LIFE FROM TABELA_COBERTURAS_POSTO WHERE ID_ITEM = "+ str(material) +" ;")

                                    material_nome= c.fetchall()
                                    title2 = title2 + "\n**Material: "+material_nome[0][0]+" | Quantidade : "+str(qnt)+" | Data : "+ str(data.day)+"-"+str(data.month)+"-"+str(data.year)+"**"+"\n"
                                    title3 = "\n"
                                    title4 = "\n"
                                    for i in filtro_relatorio:
                                        if i == 'Material':
                                            if len(title3) >= 90:
                                                title4 =title4 + " | Material: "+ str(material_nome[0][0])
                                            else:
                                                title3 =title3+ " | Material: "+ str(material_nome[0][0])
                                        if  i == "Paciente":
                                            if len(title3) >= 90:
                                                title4 =title4 + " | Paciente : "+ str(paciente_nome[0][0])
                                            else:
                                                title3 =title3 + " | Paciente : "+ str(paciente_nome[0][0])

                                        if i == "Local da lesão":
                                            if len(title3)>= 90:
                                                title4 =title4 +" | Local da lesão: "+str(local)
                                            else:
                                                title3 =title3 +" | Local da lesão: "+str(local)

                                        if i == "Quantidade":
                                            if len(title3)>= 90:
                                                title4 =title4+ " | Quantidade : "+str(qnt)
                                            else:
                                                title3 =title3+ " | Quantidade : "+str(qnt)

                                        if i == "Codigo arqlife":
                                            if len(title3)>= 90:
                                                title4 =title4+ " | Codigo arqlife : "+str(material_nome[0][1])
                                            else:
                                                title3 =title3+ " | Codigo arqlife : "+str(material_nome[0][1])

                                    pdf.cell(200, 5, txt=title3, ln=1, align="L")
                                    pdf.ln(1)
                                    pdf.cell(200, 5, txt=title4, ln=1, align="L")
                                    pdf.ln(1)
                                    pdf.cell(200, 5, txt="-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------", ln=1, align="L")
                                    pdf.ln(2)
                                    dateNow =datetime.now()
                                    if data.year == dateNow.year:
                                        list_months_verify.append(data.month)
                                if str(DataE) == "None":
                                    title69 = "\n**● Local da lesão: "+str(local)+" | Etiologia: "+str(etiologia)+" | Data de abertura : "+ str(dataA.day)+"-"+str(dataA.month)+"-"+str(dataA.year)+ " | Paciente : "+ str(paciente_nome[0][0])+" | CNS : "+ str(paciente_nome[0][1])+"**\n " +"\n"+title2
                                    if len(list_months_verify) == 0:
                                        st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" não teve pedidos esse mês")
                                    else:
                                        if max(list_months_verify) != dateNow.month:
                                                st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" não teve pedidos esse mês")

                                    for r in list3_months_verify:
                                                st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" existente a mais de 120 dias")

                                    st.subheader("-------------------------------------------------------------------------------\n"+title69)
            st.divider()
            pdf.output("example.pdf")
            with open("example.pdf", "rb") as f:
                            st.download_button(
                                label="Fazer dowload do relatorio em PDF",
                                data=f,
                                file_name='relatorio.pdf'
                            )
        else:   
            for k in list_tables:

                c.execute(
                "SELECT UNIDADE_NOME FROM TABELA_LOGINS_POSTO WHERE ID_UNIDADE = "+str(k)+" ;")
                nome= c.fetchall()
                st.divider()
                
                
                c.execute(
                    "SELECT * FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = "+ str(k) +" AND STATUS = 'ATIVO';")

                pacientes_unidades = c.fetchall()
                
                title   = "**Pacientes da unidade "+str(nome[0][0])+":**\n "


                for i in pacientes_unidades:
                    Nome = i[1]
                    Idade = i[3]
                    Genero = i[4]
                    Cns = i[5]
                    Status = i[7]
                    
                    title = title + "\n**● Nome do paciente: "+Nome+" | Idade paciente: "+str(Idade)+" | Gênero paciente : "+Genero+" | Cns do paciente : "+ str(Cns)+"**\n "
                st.subheader(title)
                st.divider()
                c.execute(
                    "SELECT * FROM TABELA_PROCESSOS_POSTO WHERE ID_UNIDADE = "+ str(k) +" AND DATA_DE_ENCERRAMENTO is null;")

                processos_unidades = c.fetchall()
                title   = "**Processos da unidade "+str(nome[0][0])+":**\n "
                titleUtil = title.replace("*","")
                pdf.cell(200, 10, txt=titleUtil.replace("Processos","Pedidos"), ln=1, align="C")
                pdf.ln(2)
                for i in processos_unidades:
                    id_processo = i[0]
                    local = i[1]
                    etiologia = i[2]
                    dataA = i[5]
                    DataE = i[6]
                    Paciente = i[8]
                    if dateNow > ((dataA + timedelta(days=120))):
                        list3_months_verify.append(data.month)
                    c.execute(
                    "SELECT NOME,CNS FROM TABELA_PACIENTES_POSTO WHERE ID_PACIENTE = "+ str(Paciente) +" ;")

                    paciente_nome= c.fetchall()

                    # Criar um dicionário vazio
                    meses = {}

                    # Adicionar os meses e os números como pares de chave-valor
                    meses["Janeiro"] = 1
                    meses["Fevereiro"] = 2
                    meses["Março"] = 3
                    meses["Abril"] = 4
                    meses["Maio"] = 5
                    meses["Junho"] = 6
                    meses["Julho"] = 7
                    meses["Agosto"] = 8
                    meses["Setembro"] = 9
                    meses["Outubro"] = 10
                    meses["Novembro"] = 11
                    meses["Dezembro"] = 12

                    c.execute(
                    "SELECT * FROM TABELA_PEDIDOS_COBERTURAS_POSTO WHERE ID_PROCESSO = "+ str(id_processo) +" AND MONTH(DATA_ATUAL) ="+str(meses[filtro_mes])+" AND YEAR(DATA_ATUAL) ="+str(filtro_ano)+" ;")

                    pedidios_processos = c.fetchall()
                    title2   = "**-Pedidos do processo:**\n "
                    for i in pedidios_processos:
                        material = i[1]
                        qnt = i[3]
                        data = i[6]
                        c.execute(
                        "SELECT MODELO_NOME,COD_ARQ_LIFE FROM TABELA_COBERTURAS_POSTO WHERE ID_ITEM = "+ str(material) +" ;")

                        material_nome= c.fetchall()
                        if material_nome[0][0] in list(dict_total_material) :
                            dict_total_material[material_nome[0][0]] = dict_total_material[material_nome[0][0]] + qnt
                        else:
                            dict_total_material[material_nome[0][0]] = qnt
                        title2 = title2 + "\n**Material: "+material_nome[0][0]+" | Quantidade : "+str(qnt)+" | Data : "+ str(data.day)+"-"+str(data.month)+"-"+str(data.year)+"**"+"\n"
                        title3 = "\n"
                        title4 = "\n"
                        for i in filtro_relatorio:
                            if i == 'Material':
                                if len(title3) >= 90:
                                    title4 =title4 + " | Material: "+ str(material_nome[0][0])
                                else:
                                    title3 =title3+ " | Material: "+ str(material_nome[0][0])
                            if  i == "Paciente":
                                if len(title3) >= 90:
                                    title4 =title4 + " | Paciente : "+ str(paciente_nome[0][0])
                                else:
                                    title3 =title3 + " | Paciente : "+ str(paciente_nome[0][0])

                            if i == "Local da lesão":
                                if len(title3)>= 90:
                                    title4 =title4 +" | Local da lesão: "+str(local)
                                else:
                                    title3 =title3 +" | Local da lesão: "+str(local)

                            if i == "Quantidade":
                                if len(title3)>= 90:
                                    title4 =title4+ " | Quantidade : "+str(qnt)
                                else:
                                    title3 =title3+ " | Quantidade : "+str(qnt)

                            if i == "Codigo arqlife":
                                if len(title3)>= 90:
                                    title4 =title4+ " | Codigo arqlife : "+str(material_nome[0][1])
                                else:
                                    title3 =title3+ " | Codigo arqlife : "+str(material_nome[0][1])

                        pdf.cell(200, 5, txt=title3, ln=1, align="L")
                        pdf.ln(1)
                        pdf.cell(200, 5, txt=title4, ln=1, align="L")
                        pdf.ln(1)
                        pdf.cell(200, 5, txt="-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------", ln=1, align="L")
                        pdf.ln(2)
                        dateNow =datetime.now()
                        if data.year == dateNow.year:
                            list_months_verify.append(data.month)
                        
                    if str(DataE) == "None":
                        title = title + "\n**● Local da lesão: "+str(local)+" | Etiologia: "+str(etiologia)+" | Data de abertura : "+ str(dataA.day)+"-"+str(dataA.month)+"-"+str(dataA.year)+ " | Paciente : "+ str(paciente_nome[0][0])+" | CNS : "+ str(paciente_nome[0][1])+"**\n " +"\n"+title2
                        if len(list_months_verify) == 0:
                            st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" não teve pedidos esse mês")
                        else:
                            if max(list_months_verify) != dateNow.month:
                                    st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" não teve pedidos esse mês")

                for r in list3_months_verify:
                        st.warning("Processo de "+str(paciente_nome[0][0])+" com cns de :"+str(paciente_nome[0][1])+" existente a mais de 120 dias")
                st.subheader(title)
            st.divider()
            for i in dict_total_material:
                    pdf.set_font("Arial", size=10)
                    pdf.cell(200,7, txt=str("Total de "+i+" é " +str(dict_total_material[i])), ln=1, align="L")
            pdf.output("example.pdf")
            with open("example.pdf", "rb") as f:
                            st.download_button(
                                label="Fazer dowload do relatorio em PDF",
                                data=f,
                                file_name='relatorio.pdf'
                            )

        ##ABA MANIPULADOR DE DADOS##
    if selected == "Manipulador de dados":

        ##CRIANDO VARIAVEIS DA SESSÃO##
        if 'new_form' not in st.session_state:
            st.session_state['new_form'] = 0

        if 'new_form2' not in st.session_state:
            st.session_state['new_form2'] = 0

        if 'new_form3' not in st.session_state:
            st.session_state['new_form3'] = " "

        if 'new_form4' not in st.session_state:
            st.session_state['new_form4'] = " "

        if 'new_form5' not in st.session_state:
            st.session_state['new_form5'] = " "

        if 'columns_number' not in st.session_state:
            st.session_state['columns_number'] = 0

        if 'tableName' not in st.session_state:
            st.session_state['tableName'] = 0

        if 'list_tablesofc' not in st.session_state:
            st.session_state['list_tablesofc'] = list_tablesofc = []

        if 'list_tablesdel' not in st.session_state:
            st.session_state['list_tablesdel'] = []

        if 'list_tablesdel2' not in st.session_state:
            st.session_state['list_tablesdel2'] = []

        if 'list_tablesdelAll' not in st.session_state:
            st.session_state['list_tablesdelAll'] = 0

        if 'list_tablesdownloadAll' not in st.session_state:
            st.session_state['list_tablesdownloadAll'] = 0

        if 'Table_orders_download' not in st.session_state:
            st.session_state['Table_orders_download'] = 0

        if 'list_tablesalter' not in st.session_state:
            st.session_state['list_tablesalter'] = []

        if 'datau' not in st.session_state:
            st.session_state['datau'] = lista_datau = []

        ##PRIMEIRA TELA#
        logout = st.button("Logout")
        if(logout):
            LogOut()
        st.divider()
        st.title("Manipulador de dados")

        if st.session_state.columns_number > 0:
            st.session_state.new_form = st.session_state.columns_number

        if st.session_state.list_tablesofc != []:
            st.session_state.new_form2 = 1

        if st.session_state.list_tablesdel != []:
            st.session_state.new_form3 = st.session_state.list_tablesdel[0]

        if st.session_state.list_tablesdel2 != []:
            st.session_state.new_form5 = st.session_state.list_tablesdel2[0]

        if st.session_state.list_tablesalter != []:
            st.session_state.new_form4 = st.session_state.list_tablesalter[0]

        if st.session_state.new_form2 > 0:

            count = 0

            for i in st.session_state.list_tablesofc:
                count = count+1
                data = pd.read_sql("SELECT * FROM "+i, con=connection)
                st.session_state.datau.append(data)
                st.subheader(str(count)+"º tabelas selecionada")
                st.dataframe(data)

            conc = pd.concat(st.session_state.datau, axis=0)
            st.subheader("Tabela concatenada")
            st.dataframe(conc)

            st.download_button(
                label="Fazer dowload da tabela concatenada",
                data=to_excel(conc),
                file_name='large_df.xlsx'
            )
            st.session_state.datau = []
            cancel_b = st.button("Cancelar")
            if cancel_b:
                st.session_state.datau = []

            st.session_state.new_form2 = 0

        elif st.session_state.new_form > 0:

            count = -1
            st.session_state.tableName = st.session_state.tableName
            st.write(st.session_state.tableName)
            st.session_state['list_columnsT'] = list_columnsT = []
            st.session_state['list_columnsN'] = list_columnsN = []

            for i in range(0, st.session_state.new_form):
                st.session_state.list_columnsN.append(str(i))
                st.session_state.list_columnsT.append(str(i))

            for i in st.session_state.list_columnsN:
                count = count+1
                count_str = int(i)+1
                list_columnsN[count] = st.text_input(
                    "Insira o nome da "+str(count_str)+"º coluna")

                list_columnsT[count] = st.selectbox(
                    "Selecione o tipo da "+str(count_str)+"º coluna", ('Numerico', 'Categorico', 'Data'))

            for i in range(0, int(len(st.session_state.list_columnsT))):
                if list_columnsT[i] == 'Numerico':
                    list_columnsT[i] = list_columnsT[i].replace(
                        'Numerico', 'int')

                if list_columnsT[i] == 'Categorico':
                    list_columnsT[i] = list_columnsT[i].replace(
                        'Categorico', 'varchar(150)')

                if list_columnsT[i] == 'Data':
                    list_columnsT[i] = list_columnsT[i].replace('Data', 'date')

            create = st.button("Criar")
            cancel_b = st.button("Cancelar")
            if cancel_b:
                st.session_state.new_form = 0

            if create:
                create_command = "CREATE TABLE " + \
                    Clean_Names(str(st.session_state.tableName))+" ("

                for i in range(0, int(len(st.session_state.list_columnsN))):
                    if list_columnsN[i] == list_columnsN[-1]:
                        create_command = create_command + \
                            Clean_Names(str(list_columnsN[i]))+" " + \
                            Clean_Names(str(list_columnsT[i]))+");"
                    else:
                        create_command = create_command + \
                            Clean_Names(str(list_columnsN[i])) + " " + \
                            Clean_Names(str(list_columnsT[i]))+","
                st.write(":green[TIPO CRIADO COM SUCESSO!]")
                c.execute(create_command)
                st.button("Continuar")
                count = -1
                st.session_state.new_form = 0

        elif st.session_state.new_form3 != " ":

            st.subheader("Você tem certeza que quer deletar o tipo de tabela " +
                         st.session_state.new_form3+"?")

            comfirmation = st.button("Sim, quero deletar")

            if comfirmation:
                c.execute("DROP TABLE "+st.session_state.new_form3)
                st.write(":green[TIPO DE TABELA DELETADO COM SUCESSO!]")
                st.session_state.new_form3 = " "
                st.button("Continuar")

            nop = st.button("Não")
            if nop:
                st.session_state.new_form3 = " "

            st.warning(
                "Cuidado ao concordar a tabela sera deletada imediatamente")

        elif st.session_state.new_form4 != " ":
            count = -1
            selec = st.radio("Selecione o tipo de alteração",
                             ('Renomear', 'Alterar Colunas'))

            if selec == 'Alterar Colunas':
                st.subheader("Selecione as colunas que você deseja alterar da tabela " +
                             st.session_state.new_form4)
                list_features = []
                c.execute(
                    "SELECT COLUMN_NAME from INFORMATION_SCHEMA.COLUMNS where table_schema = 'database' and table_name = '" + st.session_state.new_form4+"';")
                columns = c.fetchall()
                for i in columns:
                    value = i[0]
                    list_features.append(value)
                type_columns = st.multiselect('Escolha as colunas a serem alteradas',
                                              list_features)

                st.session_state['list_columnsT'] = list_columnsT = []
                st.session_state['list_columnsN'] = list_columnsN = []

                for i in range(0, len(type_columns)):
                    st.session_state.list_columnsN.append(str(type_columns[i]))
                    st.session_state.list_columnsT.append(str(type_columns[i]))

                for i in st.session_state.list_columnsN:
                    count = count+1
                    count_str = i
                    list_columnsN[count] = st.text_input(
                        "Insira o novo nome da coluna "+str(count_str))

                    list_columnsT[count] = st.selectbox(
                        "Selecione o novo tipo da coluna "+str(count_str), ('Numerico', 'Categorico', 'Data'))

                    countT = -1
                    for i in list_columnsT:
                        countT = countT + 1
                        if i == 'Numerico':
                            list_columnsT[countT] = "int"
                        if i == 'Categorico':
                            list_columnsT[countT] = "varchar(150)"
                        if i == 'Data':
                            list_columnsT[countT] = "date"

                comfirmation = st.button("Atualizar")

                if comfirmation:
                    for i in range(0, len(list_columnsN)):

                        c.execute("ALTER TABLE " +
                                  st.session_state.new_form4+" MODIFY COLUMN "+type_columns[i]+" "+list_columnsT[i])

                        c.execute("ALTER TABLE " +
                                  st.session_state.new_form4+" RENAME COLUMN "+type_columns[i]+" TO "+list_columnsN[i])

                    st.write(":green[TABELA ATUALIZADA COM SUCESSO!]")
                    st.session_state.new_form4 = " "

                    st.button("Continuar")
                nop = st.button("Não")
                if nop:
                    st.session_state.new_form4 = " "

            if selec == 'Renomear':
                st.subheader("Digite como você deseja renomear a tabela " +
                             st.session_state.new_form4)
                new_name = st.text_input("Digite o novo nome")
                ren = st.button("Renomear")

                if ren:
                    st.write()
                    c.execute("RENAME TABLE " +
                              st.session_state.new_form4+" TO "+new_name)
                    st.write(":green[TABELA RENOMEADA COM SUCESSO!]")
                    st.session_state.new_form4 = " "
                    st.button("Continuar")

                nop = st.button("Não")
                if nop:
                    st.session_state.new_form4 = " "

        elif st.session_state.new_form5 != " ":

            st.subheader("Você tem certeza que quer deletar a tabela " +
                         st.session_state.new_form5+"?")

            comfirmation = st.button("Sim, quero deletar")

            if comfirmation:
                c.execute("DROP TABLE "+st.session_state.new_form5)
                st.write(":green[TABELA DELETADA COM SUCESSO!]")
                st.session_state.new_form5 = " "
                st.button("Continuar")

            nop = st.button("Não")
            if nop:
                st.session_state.new_form5 = " "

            st.warning(
                "Cuidado ao concordar a tabela sera deletada imediatamente")

        elif st.session_state.list_tablesdelAll == 1:
            st.subheader("Tem certeza que quer deletar todas as tabelas?")
            c.execute(
                "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

            list_tables = []
            tables = c.fetchall()
            for i in tables:
                value = i[2]
                if ((('tipo' in value) | ('TIPO' in value)) & (('1' in value) | ('2' in value) | ('3' in value) | ('4' in value) | ('5' in value) | ('6' in value) | ('7' in value) | ('8' in value) | ('9' in value))):
                    list_tables.append(value)

            confirmDeleteAll = st.button("Confirmar")
            if confirmDeleteAll:
                st.subheader("Isso pode demorar um pouquinho")
                for i in list_tables:
                    with st.spinner('Aguarde...'):
                        c.execute("DROP TABLE "+i+";")
                        connection.commit()

                st.success(":green[TODAS TABELAS DELETADAS COM SUCESSO!]")

            noDeleteAll = st.button("Cancelar")
            if noDeleteAll:
                st.session_state.list_tablesdelAll = 0
                st.experimental_rerun()

        elif st.session_state.list_tablesdownloadAll == 1:
            st.subheader("Tem certeza que quer baixar todas as tabelas?")
            c.execute(
                "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

            list_tables = []
            tables = c.fetchall()
            for i in tables:
                value = i[2]
                if ((('tipo' in value) | ('TIPO' in value)) & (('1' in value) | ('2' in value) | ('3' in value) | ('4' in value) | ('5' in value) | ('6' in value) | ('7' in value) | ('8' in value) | ('9' in value))):
                    list_tables.append(value)

            confirmDownloadAll = st.button("Confirmar")
            # confirmDeleteAll = True
            if confirmDownloadAll:

                st.subheader("Isso pode demorar um pouquinho")
                zip_filename = "my_zip_file.zip"
                import zipfile
                with BytesIO() as buffer:
                    with zipfile.ZipFile(buffer, "w") as zip:
                        for i in list_tables:
                            with st.spinner('Aguarde...'):
                                data = pd.read_sql(
                                    "SELECT * FROM " + i, con=connection)
                            excel_data = to_excel(data)
                            zip.writestr(i + ".xlsx", excel_data)
                    buffer.seek(0)

                    down = st.download_button(
                        label="Fazer dowload da tabela concatenada",
                        data=buffer,
                        file_name="todasTabelas.zip"
                    )
                    st.success(":green[TODAS TABELAS BAIXADAS COM SUCESSO!]")

            noDownloadAll = st.button("Cancelar")
            if noDownloadAll:
                st.session_state.list_tablesdownloadAll = 0
                st.experimental_rerun()

        elif st.session_state.Table_orders_download == 1:

            st.subheader("Fazer download de tabela de pedidos")
            mes = st.selectbox("Selecione o mês minimo", [
                "JANEIRO",
                "FEVEREIRO",
                "MARÇO",
                "ABRIL",
                "MAIO",
                "JUNHO",
                "JULHO",
                "AGOSTO",
                "SETEMBRO",
                "OUTUBRO",
                "NOVEMBRO",
                "DEZEMBRO"])
            anos = []
            anos.append(dt.date.today().year)
            if dt.date.today().year > 2023:
                num_anos = dt.date.today().year - 2023
                for i in range(num_anos):
                    days = float(365*(i+1))
                    previous_year = dt.date.today() - timedelta(days=float(days))
                    anos.append(previous_year.year)

            ano = st.selectbox("Selecione o ano minimo", anos)
            dias = []
            for i in range(calendar.monthrange(int(ano), int(meses_dict[mes]))[1]):
                dias.append(i+1)
            dia = st.selectbox("Selecione o dia minimo", dias)

            mes2 = st.selectbox("Selecione o mês maximo", [
                "JANEIRO",
                "FEVEREIRO",
                "MARÇO",
                "ABRIL",
                "MAIO",
                "JUNHO",
                "JULHO",
                "AGOSTO",
                "SETEMBRO",
                "OUTUBRO",
                "NOVEMBRO",
                "DEZEMBRO"])
            ano2 = st.selectbox("Selecione o ano maximo", anos)

            dias2 = []
            for i in range(calendar.monthrange(int(ano2), int(meses_dict[mes2]))[1]):
                dias2.append(i+1)
            dia2 = st.selectbox("Selecione o dia maximo", dias2)
            cancel = st.button("Cancelar")
            if cancel:
                st.session_state['Table_orders_download'] = 0
                st.experimental_rerun()

        ##SEGUNDA TELA DA ABA##
        else:
            ##CONCATENAÇÃO DE TABELAS##
            new_concat = st.button("Criar concatenação")
            if new_concat:
                with st.form(key='concat_columns'):
                    c.execute(
                        "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

                    list_tables = []
                    tables = c.fetchall()
                    for i in tables:
                        value = i[2]
                        if ((('tipo' in value) | ('TIPO' in value)) & (('1' in value) | ('2' in value) | ('3' in value) | ('4' in value) | ('5' in value) | ('6' in value) | ('7' in value) | ('8' in value) | ('9' in value))):
                            list_tables.append(value)

                    list_tablesofc = st.multiselect('Escolha as tabelas a serem concatenadas',
                                                    list_tables, key='list_tablesofc')

                    submitted = st.form_submit_button(label="Enviar")
                    if submitted:
                        st.dataframe(dados, key="concat_columns")
                st.button("Cancelar")

            ##CRIANDO NOVO TIPO DE TABELA##
            new_table = st.button("Criar novo tipo de tabela")
            if new_table:
                if st.session_state.new_form == 0:

                    with st.form(key='number_columns'):
                        columnsName = st.text_input(
                            "Insira o nome da nova tabela", key='tableName')
                        columnsN = st.number_input(
                            "Insira a quntidade de colunas", min_value=0, max_value=30, key='columns_number')

                        submitted = st.form_submit_button(label="Enviar")
                    st.button("Cancelar")
            ##ALTERANDO TIPO DE TABELA##
            alter_table = st.button("Alterar tipo de tabela")
            if alter_table:
                with st.form(key='alter_columns'):
                    c.execute(
                        "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

                    list_tables = []
                    tables = c.fetchall()
                    for i in tables:
                        value = i[2]
                        if (('TIPO' in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)) | (('tipo' in value) & ('1' not in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)):
                            list_tables.append(value)

                    list_tablesdel = st.multiselect('Escolha a tabela a ser deletada',
                                                    list_tables, key='list_tablesalter', max_selections=1)

                    submitted = st.form_submit_button(label="Alterar")
                st.button("Cancelar")
            ##DELETANDO TIPO DE TABELA##
            delete_table = st.button("Deletar tipo de tabela")
            if delete_table:
                with st.form(key='delete_columns'):
                    c.execute(
                        "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

                    list_tables = []
                    tables = c.fetchall()
                    for i in tables:
                        value = i[2]
                        if (('TIPO' in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)) | (('tipo' in value) & ('1' not in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)):
                            list_tables.append(value)

                    list_tablesdel = st.multiselect('Escolha a tipo de tabela a ser deletada',
                                                    list_tables, key='list_tablesdel', max_selections=1)

                    submitted = st.form_submit_button(label="Deletar")
                st.button("Cancelar")

            ##DELETANDO TABELA##
            delete_table2 = st.button("Deletar tabela")
            if delete_table2:
                with st.form(key='delete_columns2'):
                    c.execute(
                        "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

                    list_tables = []
                    tables = c.fetchall()
                    for i in tables:
                        value = i[2]
                        if ((('tipo' in value) | ('TIPO' in value)) & (('1' in value) | ('2' in value) | ('3' in value) | ('4' in value) | ('5' in value) | ('6' in value) | ('7' in value) | ('8' in value) | ('9' in value))):
                            list_tables.append(value)

                    list_tablesdel = st.multiselect('Escolha a tabela a ser deletada',
                                                    list_tables, key='list_tablesdel2', max_selections=1)

                    submitted = st.form_submit_button(label="Deletar")
                st.button("Cancelar")
            ##DELETANDO TODAS TABELAS##
            delete_Alltable = st.button("Deletar todas tabelas")
            if delete_Alltable:
                st.session_state.list_tablesdelAll = 1
                st.experimental_rerun()

            ##BAIXANDO TODAS TABELAS##
            download_Alltable = st.button("Baixar todas tabelas")
            if download_Alltable:
                st.session_state.list_tablesdownloadAll = 1
                st.experimental_rerun()

            baixar_pedidos = st.button("Baixar pedidos")

            if baixar_pedidos:
                st.session_state.Table_orders_download = 1
                st.experimental_rerun()

        ##ABA ANALISE DE DADOS##
    if selected == "Analise de dados":
        logout = st.button("Logout")
        if(logout):
            LogOut()
        st.divider()
        st.title("Analise de dados")
        select_vertentes = st.selectbox("Selecione as vertentes que quer analisar", [
            'PEDIDOS', 'PACIENTES', 'PROCESSOS'])


        list_tables = []
        c.execute(
            "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")

        tables = c.fetchall()
        for i in tables:
            value = i[2]
            if (('POSTO' in value) & (select_vertentes in value)):
                list_tables.append(value)

        list_features = []
        for i in list_tables:
            c.execute(
                "SELECT COLUMN_NAME from INFORMATION_SCHEMA.COLUMNS where table_schema = 'database' and table_name = '" + i+"';")
            list_features = []
            columns = c.fetchall()
            for x in columns:
                value = x[0]
                list_features.append(value)

            lista_features_str = []
            for i in list_features:
                i = i.replace("_ID","")
                i = i.replace("ID_","")
                lista_features_str.append(i)
            
            newVar = []
            var = st.multiselect(
                "Selicione as variaveis que deseja analisar da vertente ", lista_features_str)
            for x in var:
                for i in list_features:
                    if x in i:
                        newVar.append(i)

            
        count_index_features = []
        count_count_index_features = -1
        another_count = len(var)

        list_types = []
        for i in newVar:
            c.execute(
                "SELECT DATA_TYPE  FROM INFORMATION_SCHEMA.COLUMNS  WHERE table_name ='"+list_tables[0] + "' AND column_name = '"+i+"';")

            types = c.fetchall()
            for x in types:
                if "ID" in i:
                    value = "varchar"
                    list_types.append(value)
                else:
                    value = x[0]
                    list_types.append(value)

        count_var = -1
        for i in list_types:
            count_var = count_var+1
            # st.write(i)
            list_features_unique = []
            if i == 'varchar':
                # st.write(count_var)
                if newVar[count_var] == "ID_MATERIAL":

                    c.execute("SELECT DISTINCT MODELO_NOME FROM TABELA_COBERTURAS_POSTO ;")
                    unique_var = c.fetchall()
                    list_features_unique = []
                    for x in unique_var:
                        value = x[0]
                        list_features_unique.append(value)
                        
                elif newVar[count_var] == "PACIENTE_ID": 
   
                    c.execute("SELECT DISTINCT NOME FROM TABELA_PACIENTES_POSTO ;")
                    unique_var = c.fetchall()
                    list_features_unique = []
                    for x in unique_var:
                        value = x[0]
                        list_features_unique.append(value)
                else:

                    c.execute("SELECT DISTINCT " +
                    newVar[count_var]+" FROM "+list_tables[0]+";")
                    unique_var = c.fetchall()
                    list_features_unique = []
                    for x in unique_var:
                        value = x[0]
                        list_features_unique.append(value)

            if i == 'datetime':

                # st.write(list_features)
                # st.write("SELECT DISTINCT " +
                # list_features[count_var]+" FROM "+list_tables[0]+";")
                c.execute("SELECT DISTINCT " +
                newVar[count_var]+" FROM "+list_tables[0]+";")
                unique_var = c.fetchall()
                list_features_unique = []
                for x in unique_var:
                    value = x[0]
                    list_features_unique.append(value)

            if i == 'int':
                c.execute("SELECT DISTINCT " +
                newVar[count_var]+" FROM "+list_tables[0]+";")
                unique_var = c.fetchall()
                list_features_unique = []
                for x in unique_var:
                    value = x[0]
                    list_features_unique.append(value)
            list_value_selected = st.selectbox(
                    "Selecione o valor do pelo qual quer filtrar " +var[count_var] , list_features_unique)


        ##ABA SUBIR TABELAS##
    if ((selected == ("Subir tabelas"))):
        logout = st.button("Logout")
        if(logout):
            LogOut()
        st.divider()

        if((dt.date.today().day >= 20) & (dt.date.today().day <= 25)):
            st.session_state.new_form2 = 0
            st.divider()
            st.title("Insira sua tabela e as informações necessarias abaixo")
            c.execute(
                "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")
            list_tables = []

            tables = c.fetchall()
            for i in tables:
                value = i[2]
                if (('TIPO' in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)) | (('tipo' in value) & ('1' not in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)):
                    list_tables.append(value)

            selection_type = st.selectbox("Selecione o tipo da tabela",
                                          list_tables)

            st.subheader(
                ":red[Clique no botão a baixo para subir a tabela ⇩]")

            c.execute(
                "SELECT COLUMN_NAME from INFORMATION_SCHEMA.COLUMNS where table_schema = 'database' and table_name = '"+selection_type+"'")
            columns = c.fetchall()
            css = '''
                    <style>
                    [data-testid="stFileUploadDropzone"] div div::before {color:green; content:"Clique aqui para subir sua tabela"}
                    [data-testid="stFileUploadDropzone"] div div span{display:none;}
                    [data-testid="stFileUploadDropzone"] div div::after { font-size: .8em; content:""}
                    [data-testid="stFileUploadDropzone"] div div small{display:none;}
                    [data-testid="stFileUploadDropzone"] button[kind="secondary"]::before {
                    content: "Carregar tabela /  ";
                    }
                    [data-testid="stFileUploadDropzone"] button[kind="secondary"]{visibility: hidden;}
                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            dados4 = st.file_uploader("Tabela", type=["xlsx"])

            if (dados4 != None) & (selection_type == "MAPA_MENSAL_COMISSAO_TIPO"):
                dados4 = pd.read_excel(
                    dados4, sheet_name='BASE DE DADOS', engine='openpyxl')
                dados4 = dados4.drop([0, 1, 2, 3, 4], axis=0)
                dados4.columns = dados4.iloc[0].values
                dados4 = dados4.drop(5, axis=0)
                dados4 = dados4.reset_index()
                dados4 = dados4.drop('index', axis=1)
                for i in dados4.columns:
                    dados4 = dados4.rename({i: Clean_Names(i)}, axis=1)

                number_columns_verify = []
                for i in dados4.columns:
                    for x in columns:
                        if i in x:
                            number_columns_verify.append(x)

                if(dados4.shape[1] == len(columns)) & (len(number_columns_verify) == dados4.shape[1]):

                    st.dataframe(dados4)
                    name = st.text_input("Nome da unidade")
                    date = st.text_input("Data do envio da tabela")
                    st.warning(
                        "LEMBRE-SE DE INSERIR O NOME DA TABELA TODO EM MAIUSCULO,SEM DIGITOS,SEM ACENTUAÇÃO E COM '_' NO LUGAR DOS ESPAÇOS, EXEMPLO: POSTO_UM  ")
                    st.warning(
                        "LEMBRE-SE DE INSERIR A DATA DA TABELA COM '_' NO LUGAR DOS ESPAÇOS, EXEMPLO: 24_06_2023  ")
                    if ((name == " ") or (date == " ") or ('/' in date) or ('-' in date) or ('?' in name) or ('á' in name) or ('à' in name) or ('â' in name) or ('ã' in name) or ('ä' in name) or ('é' in name) or ('è' in name) or ('ê' in name) or ('ë' in name) or ('í' in name) or ('ì' in name) or ('î' in name) or ('ï' in name) or ('ó' in name) or ('ò' in name) or ('ô' in name) or ('õ' in name) or ('ö' in name) or ('ú' in name) or ('ù' in name) or ('û' in name) or ('ü' in name) or ('Á' in name) or ('À' in name) or ('Â' in name) or ('Ã' in name) or ('Ä' in name) or ('É' in name) or ('È' in name) or ('Ê' in name) or ('Ë' in name) or ('Í' in name) or ('Ì' in name) or ('Î' in name) or ('Ï' in name) or ('Ó' in name) or ('Ò' in name) or ('Ô' in name) or ('Õ' in name) or ('Ö' in name) or ('Ú' in name) or ('Ù' in name) or ('Û' in name) or ('Ü' in name) or (' ' in name or (('1' in name) | ('2' in name) | ('3' in name) | ('4' in name) | ('5' in name) | ('6' in name) | ('7' in name) | ('8' in name) | ('9' in name)))):
                        st.write(
                            ":red[DATA OU NOME COM CONFIGURAÇÃO ERRADA MUDE PARA PROSSEGUIR]")
                    else:

                        nameFinal = name+date+str(selection_type)

                        ssl_args = {'ssl_ca': "cacert-2023-01-10.pem"}

                        engine = create_engine(
                            'mysql+mysqlconnector://'+st.secrets["db_username"]+':'+st.secrets["db_password"]+'@aws.connect.psdb.cloud/database', connect_args=ssl_args)
                        send_table = st.button("Enviar Tabela")
                        if send_table:
                            dados4.to_sql(nameFinal, con=engine,
                                          if_exists='replace', index=False)
                            st.write("Tabela enviada com sucesso!")
                else:
                    st.warning("Tipo não compatível")

            elif (dados4 != None):
                dados4 = pd.read_excel(dados4)
                for i in dados4.columns:
                    dados4 = dados4.rename({i: Clean_Names(i)}, axis=1)

                number_columns_verify = []
                for i in dados4.columns:
                    for x in columns:
                        if i in x:
                            number_columns_verify.append(x)

                if(dados4.shape[1] == len(columns)) & (len(number_columns_verify) == dados4.shape[1]):

                    st.dataframe(dados4)
                    name = st.text_input("Nome da unidade")
                    date = st.text_input("Data do envio da tabela")
                    st.warning(
                        "LEMBRE-SE DE INSERIR O NOME DA TABELA TODO EM MAIUSCULO E SEM NUMEROS COM A PALAVRA TIPO E _ NO LUGAR DOS ESPAÇÕS")
                    nameFinal = name+date+str(selection_type)

                    ssl_args = {'ssl': "cacert-2023-01-10.pem"}

                    engine = create_engine(
                        'mysql+mysqldb://'+st.secrets["db_username"]+':'+st.secrets["db_password"]+'@aws.connect.psdb.cloud/database', connect_args=ssl_args)
                    send_table = st.button("Enviar Tabela")
                    if send_table:
                        dados4.to_sql(nameFinal, con=engine,
                                      if_exists='replace', index=False)
                else:
                    st.error("Tipo não compatível")
        else:
            st.divider()
            st.error(
                "Acesso negado a função de subir tabelas pois esta fora da data permitidaa")
            st.warning(
                "Somente entre  os dias 20 e 25 é permitido subir tabela")
            
    if 'menu_gerenciar_logins' not in st.session_state:
        st.session_state['menu_gerenciar_logins'] = 0
        

    if ((selected == ("Gerenciar Logins"))):
        logout = st.button("Logout")
        if(logout):
                LogOut()
        st.divider()
        st.title("Gerenciar Logins")
        if st.session_state.menu_gerenciar_logins == 0:

            CreateLogin = st.button("Criar login")
            if CreateLogin:
                 st.session_state.menu_gerenciar_logins = 1
                 st.experimental_rerun()

            DeleteLogin = st.button("Excluir login")
            if DeleteLogin:
                st.session_state.menu_gerenciar_logins = 2
                st.experimental_rerun()

            TransferPacient = st.button("Transferir pacientes entre unidades")
            if TransferPacient:
                st.session_state.menu_gerenciar_logins = 4
                st.experimental_rerun()

            AddCobertura = st.button("Adicionar cobertura")
            if AddCobertura:
                st.session_state.menu_gerenciar_logins = 5
                st.experimental_rerun()

            AddPartes = st.button("Adicionar partes do corpo")
            if AddPartes:
                st.session_state.menu_gerenciar_logins = 6
                st.experimental_rerun()
            
            AddLocal = st.button("Adicionar local da lesão")
            if AddLocal:
                st.session_state.menu_gerenciar_logins = 7
                st.experimental_rerun()

            DeleteCobertura = st.button("Deletar cobertura")
            if DeleteCobertura:
                st.session_state.menu_gerenciar_logins = 8
                st.experimental_rerun()

            DeletePartes = st.button("Deletar partes do corpo")
            if DeletePartes:
                st.session_state.menu_gerenciar_logins = 9
                st.experimental_rerun()
            
            DeleteLocal = st.button("Deletar local da lesão")
            if DeleteLocal:
                st.session_state.menu_gerenciar_logins = 10
                st.experimental_rerun()
            # ChangeLogin = st.button("Mudar informações do login")
            # if ChangeLogin:
            #     st.session_state.menu_gerenciar_logins = 3
            #     st.experimental_rerun()
        
        if st.session_state.menu_gerenciar_logins == 1:
            name_confirm = False
            pass_confirm = False
            st.title("Criar login")
            login_name = st.text_input("Digite o nome do novo usuario")
            verificar_campo_texto(login_name)
            login_pass = st.text_input("Digite a senha do novo usuario",type="password")
            login_unidade = st.text_input("Digite o nome da unidade")
            if verificar_campo_texto(login_name):
                name_confirm = True
            else:
                st.error("Nome de login com caracteres proibidos")
            if len(login_name) ==0:
                name_confirm = False
                st.error("Nome de login em branco")

            if verificar_campo_texto(login_unidade):
                pass_confirm = True
            else:
                st.error("Unidade de login com caracteres proibidos")
            if len(login_unidade) ==0:
                pass_confirm = False
                st.error("Unidade de login em branco")
            crypt = bcrypt.hashpw(login_pass.encode('utf-8'), salt)
            if name_confirm & pass_confirm:
                enviar = st.button("Enviar")
                st.write()
                if enviar:
                    c.execute("INSERT INTO TABELA_LOGINS_POSTO (NIVEL_PERMISSAO,NOME_LOGIN,SENHA_LOGIN,UNIDADE_NOME)  VALUES(2,'"+login_name+"','"+str(crypt).replace("'",":")+"','"+login_unidade+"')")
                    connection.commit()
                    st.write(":green[Login cadastrado]")


            cancel = st.button("Voltar")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 3:
            st.title("Mudar login")
            st.text_input("Digite o novo nome usuario")
            st.text_input("Digite a nova senha do usuario",type="password")
            st.text_input("Digite o novo nome da unidade")
            cancel = st.button("Voltar")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 4:
            st.title("Transferir pacientes entre unidades")
            c.execute(
            "SELECT UNIDADE_NOME FROM TABELA_LOGINS_POSTO;")
            tables = c.fetchall()
            list_tables =[]
            for i in tables:
                value = i[0]
                list_tables.append(value)
            unidade = st.selectbox("Selecione a Unidade",list_tables)
            c.execute(
            "SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO WHERE UNIDADE_NOME = '"+ str(unidade) +"' ;")
            id_unidade = c.fetchall()
            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = "+ str(id_unidade[0][0]) +" AND STATUS = 'ATIVO';")
            
            list_tables =[]
            tables = c.fetchall()
            for i in tables:
                value = i[0]
                list_tables.append(value)
            paciente = st.selectbox("Selecione o paciente que deseja transferir",list_tables)

            c.execute("SELECT CNS,ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
            cns = c.fetchall()
            st.subheader("CNS do paciente: "+str(cns[0][0]))
            c.execute(
            "SELECT UNIDADE_NOME FROM TABELA_LOGINS_POSTO;")
            tables = c.fetchall()
            list_tables =[]
            for i in tables:
                value = i[0]
                list_tables.append(value)
            list_tables.remove(unidade)
            unidade = st.selectbox("Selecione a Unidade para qual vai trasnferir",list_tables)
            c.execute(
            "SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO WHERE UNIDADE_NOME = '"+ str(unidade) +"' ;")
            id_unidade = c.fetchall()


            transferencia = st.button("Transferir")
            if transferencia:
                c.execute("UPDATE TABELA_PACIENTES_POSTO SET ID_UNIDADE="+ str(id_unidade[0][0])+" WHERE CNS = "+str(cns[0][0])+" ;")
                connection.commit()
                c.execute("UPDATE TABELA_PROCESSOS_POSTO SET ID_UNIDADE="+ str(id_unidade[0][0])+" WHERE ID_PACIENTE= "+str(cns[0][1])+" ;")
                connection.commit()
                st.write(":green[Paciente trasnferido com sucesso!]")
                
            cancel = st.button("Voltar")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 2:
            st.title("Excluir login")
            c.execute("SELECT NOME_LOGIN FROM TABELA_LOGINS_POSTO")
            list_exclude_login = c.fetchall()
            list_exclude_login_total = []
            for x in list_exclude_login:
                value = x[0]
                list_exclude_login_total.append(value)

            exclude_login = st.selectbox("Selecionar login a ser deletado",list_exclude_login_total)
            c.execute("SELECT ID_UNIDADE FROM TABELA_LOGINS_POSTO WHERE NOME_LOGIN = '"+exclude_login+"' ;")
            list_exclude_login = c.fetchall()
            list_exclude_login_total = []
            for x in list_exclude_login:
                value = x[0]
                st.write()
                list_exclude_login_total.append(value)

            deletar = st.button("Deletar")
            if deletar:
                c.execute("DELETE FROM TABELA_LOGINS_POSTO WHERE ID_UNIDADE = "+str(list_exclude_login_total[0])+";")
                c.execute("INSERT INTO LAST_DELETED_POSTO (ULTIMO_DELETADO) VALUES( "+str(list_exclude_login_total[0])+" );")
                connection.commit()

                # if st.write(str(cookie_manager.get(cookie=cookiee))) == list_exclude_login_total[0]:
                #     LogOut()

                st.write(":green[Login deletado com sucesso!]")
            cancel = st.button("Voltar")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 5:
            st.title("Adicionar uma cobertura a lista de coberturas")
            name = st.text_input("Insira o nome da cobertura")
            desc = st.text_input("Insira a descrição")
            cod = st.text_input("Insira o codigo ARQ life")
            cancel = st.button("Voltar")
            enviar = st.button("Enviar")
            if enviar:
                    c.execute("INSERT INTO TABELA_COBERTURAS_POSTO (MODELO_NOME,DESCRICAO,COD_ARQ_LIFE) VALUES ('"+str(name)+"','"+str(desc)+"',"+str(cod)+");")
                    connection.commit()
                    st.write(":green[Cobertura adicionada com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()


        if st.session_state.menu_gerenciar_logins == 6:
            st.title("Adicionar uma etiologia lista de etiologias")
            name = st.text_input("Insira o nome da etiologia")
            cancel = st.button("Voltar")
            enviar = st.button("Enviar")
            if enviar:
                    c.execute("INSERT INTO TABELA_ETIOLOGIAS_POSTO (NOME_ETIOLOGIA) VALUES ('"+str(name)+"');")
                    connection.commit()
                    st.write(":green[Etiologia adicionada com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()


        if st.session_state.menu_gerenciar_logins == 7:
            st.title("Adicionar uma local da lesão a lista de locais de lesão")
            name =st.text_input("Insira o nome da local da lesão")
            enviar = st.button("Enviar")
            cancel = st.button("Voltar")
            enviar = st.button("Enviar")
            if enviar:
                    c.execute("INSERT INTO TABELA_LOCAIS_LESAO (NOME_LOCAL_LESAO) VALUES ('"+str(name)+"');")
                    connection.commit()
                    st.write(":green[Local da lesão adicionado com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 8:
            st.title("Deletar uma cobertura da lista de coberturas")
            list_names = []
            c.execute("SELECT MODELO_NOME FROM TABELA_COBERTURAS_POSTO;")
            names =c.fetchall()
            if len(names) == 0:
                st.error("Não existe nenhuma cobertura registrada ")
            for i in names:
                list_names.append(i[0])

            name =st.selectbox("Escolha a cobertura que deseja excluir",list_names)
            cancel = st.button("Voltar")
            if (len(list_names)>0):
                enviar = st.button("Enviar")
                if enviar:
                    c.execute("DELETE FROM TABELA_COBERTURAS_POSTO WHERE MODELO_NOME ='"+name+"' ;")
                    connection.commit()
                    st.write(":green[Cobertura deletada com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 9:
            st.title("Deletar uma etiologia lista de etiologias ")
            list_names = []
            c.execute("SELECT NOME_ETIOLOGIA FROM TABELA_ETIOLOGIAS_POSTO;")
            names =c.fetchall()
            if len(names) == 0:
                st.error("Não existe nenhuma etiologia registrada ")
            for i in names:
                list_names.append(i[0])

            name =st.selectbox("Escolha a etiologia que deseja excluir",list_names)
            cancel = st.button("Voltar")
            if (len(list_names)>0):
                enviar = st.button("Enviar")
                if enviar:
                    c.execute("DELETE FROM TABELA_ETIOLOGIAS_POSTO WHERE NOME_ETIOLOGIA ='"+name+"' ;")
                    connection.commit()
                    st.write(":green[Etiologia deletada com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()

        if st.session_state.menu_gerenciar_logins == 10:
            st.title("Deletar um local de lesão da lista de locais de lesão")
            list_names = []
            c.execute("SELECT NOME_LOCAL_LESAO FROM TABELA_LOCAIS_LESAO;")
            names =c.fetchall()
            if len(names) == 0:
                st.error("Não existe nenhum local de lesão registrado ")
            for i in names:
                list_names.append(i[0])

            name =st.selectbox("Escolha o local que deseja excluir",list_names)
            cancel = st.button("Voltar")
            if (len(list_names)>0):
                enviar = st.button("Enviar")
                if enviar:
                    c.execute("DELETE FROM TABELA_LOCAIS_LESAO WHERE NOME_LOCAL_LESAO ='"+name+"' ;")
                    connection.commit()
                    st.write(":green[Local da lesão deletado com sucesso!]")
            if cancel:
                st.session_state.menu_gerenciar_logins = 0
                st.experimental_rerun()
##LOGIN USER##

##CRIANDO MENU##
if(st.session_state.Login2 == 2):
    # authenticator.logout('Logout', 'main')
    with st.sidebar:
        # selected = option_menu(
        #     menu_title="Menu",
        #     options=["Subir tabelas", "Gerenciar pacientes da unidade",
        #              "Gerenciar pedidos de coberturas", "Gerenciar processos de pacientes"],
        #     menu_icon="border-width"
        # )

        selected = option_menu(
            menu_title="Menu",
            options=[ "Gerenciar pacientes da unidade",
                     "Gerenciar pedidos de coberturas", "Gerenciar processos de pacientes"],
            menu_icon="border-width"
        )
        st.sidebar.image(
            "WhatsApp Image 2023-02-21 at 14.22.25 (1).png", use_column_width=True)
    if 'new_form2' not in st.session_state:
        st.session_state['new_form2'] = 0

        ##ABA SUBIR TABELAS#
    if ((selected == ("Subir tabelas"))):
        logout = st.button("Logout")
        if(logout):
            LogOut()
        if((dt.date.today().day >= 20) & (dt.date.today().day <= 25)):
            st.session_state.new_form2 = 0
            st.divider()
            st.title("Insira sua tabela e as informações necessarias abaixo")
            c.execute(
                "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE table_schema = 'database';")
            list_tables = []

            tables = c.fetchall()
            for i in tables:
                value = i[2]
                if (('TIPO' in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)) | (('tipo' in value) & ('1' not in value) & ('1' not in value) & ('2' not in value) and ('3' not in value) and ('4' not in value) and ('5' not in value) and ('6' not in value) and ('7' not in value) and ('8' not in value) and ('9' not in value)):
                    list_tables.append(value)

            selection_type = st.selectbox("Selecione o tipo da tabela",
                                          list_tables)

            st.subheader(
                ":red[Clique no botão a baixo para subir a tabela ⇩]")

            c.execute(
                "SELECT COLUMN_NAME from INFORMATION_SCHEMA.COLUMNS where table_schema = 'database' and table_name = '"+selection_type+"'")
            columns = c.fetchall()
            css = '''
                    <style>
                    [data-testid="stFileUploadDropzone"] div div::before {color:green; content:"Clique aqui para subir sua tabela"}
                    [data-testid="stFileUploadDropzone"] div div span{display:none;}
                    [data-testid="stFileUploadDropzone"] div div::after { font-size: .8em; content:""}
                    [data-testid="stFileUploadDropzone"] div div small{display:none;}
                    [data-testid="stFileUploadDropzone"] button[kind="secondary"]::before {
                    content: "Carregar tabela /  ";
                    }
                    [data-testid="stFileUploadDropzone"] button[kind="secondary"]{visibility: hidden;}
                    </style>
                    '''

            st.markdown(css, unsafe_allow_html=True)
            dados4 = st.file_uploader("Tabela", type=["xlsx"])

            if (dados4 != None) & (selection_type == "MAPA_MENSAL_COMISSAO_TIPO"):
                dados4 = pd.read_excel(
                    dados4, sheet_name='BASE DE DADOS', engine='openpyxl')
                dados4 = dados4.drop([0, 1, 2, 3, 4], axis=0)
                dados4.columns = dados4.iloc[0].values
                dados4 = dados4.drop(5, axis=0)
                dados4 = dados4.reset_index()
                dados4 = dados4.drop('index', axis=1)
                for i in dados4.columns:
                    dados4 = dados4.rename({i: Clean_Names(i)}, axis=1)

                number_columns_verify = []
                for i in dados4.columns:
                    for x in columns:
                        if i in x:
                            number_columns_verify.append(x)

                if(dados4.shape[1] == len(columns)) & (len(number_columns_verify) == dados4.shape[1]):

                    st.dataframe(dados4)
                    name = st.text_input("Nome da unidade")
                    date = st.text_input("Data do envio da tabela")
                    st.warning(
                        "LEMBRE-SE DE INSERIR O NOME DA TABELA TODO EM MAIUSCULO,SEM DIGITOS,SEM ACENTUAÇÃO E COM '_' NO LUGAR DOS ESPAÇOS, EXEMPLO: POSTO_UM  ")
                    st.warning(
                        "LEMBRE-SE DE INSERIR A DATA DA TABELA COM '_' NO LUGAR DOS ESPAÇOS, EXEMPLO: 24_06_2023  ")
                    if ((name == " ") or (date == " ") or ('/' in date) or ('-' in date) or ('?' in name) or ('á' in name) or ('à' in name) or ('â' in name) or ('ã' in name) or ('ä' in name) or ('é' in name) or ('è' in name) or ('ê' in name) or ('ë' in name) or ('í' in name) or ('ì' in name) or ('î' in name) or ('ï' in name) or ('ó' in name) or ('ò' in name) or ('ô' in name) or ('õ' in name) or ('ö' in name) or ('ú' in name) or ('ù' in name) or ('û' in name) or ('ü' in name) or ('Á' in name) or ('À' in name) or ('Â' in name) or ('Ã' in name) or ('Ä' in name) or ('É' in name) or ('È' in name) or ('Ê' in name) or ('Ë' in name) or ('Í' in name) or ('Ì' in name) or ('Î' in name) or ('Ï' in name) or ('Ó' in name) or ('Ò' in name) or ('Ô' in name) or ('Õ' in name) or ('Ö' in name) or ('Ú' in name) or ('Ù' in name) or ('Û' in name) or ('Ü' in name) or (' ' in name or (('1' in name) | ('2' in name) | ('3' in name) | ('4' in name) | ('5' in name) | ('6' in name) | ('7' in name) | ('8' in name) | ('9' in name)))):
                        st.write(
                            ":red[DATA OU NOME COM CONFIGURAÇÃO ERRADA MUDE PARA PROSSEGUIR]")
                    else:

                        nameFinal = name+date+str(selection_type)

                        ssl_args = {'ssl_ca': "cacert-2023-01-10.pem"}

                        engine = create_engine(
                            'mysql+mysqlconnector://'+st.secrets["db_username"]+':'+st.secrets["db_password"]+'@aws.connect.psdb.cloud/database', connect_args=ssl_args)
                        send_table = st.button("Enviar Tabela")
                        if send_table:
                            dados4.to_sql(nameFinal, con=engine,
                                          if_exists='replace', index=False)
                            st.write("Tabela enviada com sucesso!")
                else:
                    st.warning("Tipo não compatível")

            elif (dados4 != None):
                dados4 = pd.read_excel(dados4)
                for i in dados4.columns:
                    dados4 = dados4.rename({i: Clean_Names(i)}, axis=1)

                number_columns_verify = []
                for i in dados4.columns:
                    for x in columns:
                        if i in x:
                            number_columns_verify.append(x)

                if(dados4.shape[1] == len(columns)) & (len(number_columns_verify) == dados4.shape[1]):

                    st.dataframe(dados4)
                    name = st.text_input("Nome da unidade")
                    date = st.text_input("Data do envio da tabela")
                    st.warning(
                        "LEMBRE-SE DE INSERIR O NOME DA TABELA TODO EM MAIUSCULO E SEM NUMEROS COM A PALAVRA TIPO E _ NO LUGAR DOS ESPAÇÕS")
                    nameFinal = name+date+str(selection_type)

                    ssl_args = {'ssl': "cacert-2023-01-10.pem"}

                    engine = create_engine(
                        'mysql+mysqldb://'+st.secrets["db_username"]+':'+st.secrets["db_password"]+'@aws.connect.psdb.cloud/database', connect_args=ssl_args)
                    send_table = st.button("Enviar Tabela")
                    if send_table:
                        dados4.to_sql(nameFinal, con=engine,
                                      if_exists='replace', index=False)
                else:
                    st.error("Tipo não compatível")
        else:
            st.divider()
            st.error(
                "Acesso negado a função de subir tabelas pois esta fora da data permitida")
            st.warning(
                "Somente entre  os dias 20 e 25 é permitido subir tabela")

    if 'state_processos' not in st.session_state:
        st.session_state['state_processos'] = 0

    if 'state_pacientes' not in st.session_state:
        st.session_state['state_pacientes'] = 0

    if 'state_pedidos' not in st.session_state:
        st.session_state['state_pedidos'] = 0

    if(selected == "Gerenciar pacientes da unidade"):
        logout = st.button("Logout")
        if(logout):
            LogOut()
        if st.session_state['state_pacientes'] == 0:
            st.divider()
            st.title("Gerenciar pacientes da unidade")
            add_paciente = st.button("Adicionar paciente")
            if add_paciente:
                st.session_state['state_pacientes'] = 1
                st.experimental_rerun()

            remove_paciente = st.button("Excluir paciente")
            if remove_paciente:
                st.session_state['state_pacientes'] = 2
                st.experimental_rerun()

            verify_paciente = st.button("Editar paciente")
            if verify_paciente:
                st.session_state['state_pacientes'] = 3
                st.experimental_rerun()

        if st.session_state['state_pacientes'] == 1:

            st.divider()
            st.title("Adicionar paciente")
            name_paciente = st.text_input("Insira o nome do paciente")
            data_paciente = st.text_input(
                "Insira a data de nascimento do paciente", max_chars=10)
            genero_paciente = st.selectbox("Selecione o gênero do paciente", [
                "Masculino", "Feminino"])

            cns_paciente = st.text_input("Insira o CNS do paciente")
            
            list_pacientes = []
            c.execute(
                "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE CNS = '" + str(cns_paciente) + "';")
            Id_selected = c.fetchall()
            for i in Id_selected:
                list_pacientes.append(i)
                
            st.warning(
                "LEMBRE-SE DE INSERIR A DATA DE NASCIMENTO COM '/' NO LUGAR DOS ESPAÇOS, EXEMPLO: 20/05/2023  ")

            
            data_paciente =  data_paciente.replace("/","-")
            if (data_paciente != "") & (verificar_formato_data(data_paciente)):
                data_nacs = datetime.strptime(data_paciente, '%d-%m-%Y')
                data_paciente = data_paciente[6:11] + \
                    data_paciente[2:7]+data_paciente[0:1]

                data_atual = datetime.now()

                idade = data_atual - data_nacs
                idade = math.floor(idade.days / 365)
            if (len(list_pacientes)>0):
                st.error("Usuario ja existe no sistema")

            if (verificar_formato_nome(name_paciente)) & ((name_paciente != "") & (data_paciente != "") & (verificar_formato_CNS(cns_paciente)) & (len(list_pacientes)==0)):
                enviar = st.button("Enviar")
                if enviar:
                    c.execute("INSERT INTO TABELA_PACIENTES_POSTO ( NOME, DATA_NASC, IDADE, GENERO, CNS, ID_UNIDADE,STATUS) VALUES ('"+name_paciente + "', STR_TO_DATE('" +
                              data_paciente+"','%Y-%m-%d')," + str(idade) + ", '" +
                              genero_paciente + "','" +
                              str(cns_paciente)+"'," +
                              str(st.session_state.ActualUser) +
                              ",'"+"ATIVO"+"');"
                              )

                    connection.commit()
                    st.write(":green[Usuario cadastrado com sucesso!]")

            cancel = st.button("Voltar")
            if cancel:
                st.session_state['state_pacientes'] = 0
                st.experimental_rerun()

        if st.session_state['state_pacientes'] == 3:
            st.divider()
            st.title("Editar paciente")
            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)
            if len(list_pacientes) ==0:
                st.error("Não foi possível encontrar pacientes")

            else:
                paciente = st.selectbox(
                    "Selecione o paciente ", list_pacientes)
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))
                list_pacientes = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                Id_selected = c.fetchall()
                for i in Id_selected:
                    list_pacientes.append(i)

                if(paciente != []):

                    name_paciente = st.text_input("Insira o nome do paciente")
                    data_paciente = st.text_input(
                        "Insira a data de nascimento do paciente", max_chars=10)
                    st.warning(
                        "LEMBRE-SE DE INSERIR A DATA DE NASCIMENTO COM '/' NO LUGAR DOS ESPAÇOS, EXEMPLO: 20/05/2023  ")

                    
                    data_paciente =  data_paciente.replace("/","-")
                    
                    list_pacientes = []
                    c.execute(
                        "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(name_paciente) + "';")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        list_pacientes.append(i)

                    genero_paciente = st.selectbox("Selecione o gênero do paciente", [
                        "Masculino", "Feminino"])

                    cns_paciente = st.text_input("Insira o CNS do paciente")
                    if (data_paciente != "") & (verificar_formato_data(data_paciente)):
                        data_nacs = datetime.strptime(data_paciente, '%d-%m-%Y')
                        data_paciente = data_paciente[6:11] + \
                            data_paciente[2:7]+data_paciente[0:1]

                        data_atual = datetime.now()
                        idade = data_atual - data_nacs
                        idade = math.floor(idade.days / 365)

                        
                    if (verificar_formato_nome(name_paciente)) & ((name_paciente != "") & (data_paciente != "") & (cns_paciente != "") & (verificar_formato_CNS(cns_paciente)) ):
                        enviar = st.button("Enviar")
                        if enviar:
                            c.execute("UPDATE TABELA_PACIENTES_POSTO SET NOME='" + name_paciente +
                                    "', DATA_NASC=STR_TO_DATE('" + data_paciente + "', '%Y-%m-%d'), IDADE=" +
                                    str(idade) + ", GENERO='" + genero_paciente + "', CNS='" +
                                    str(cns_paciente) + "' WHERE ID_PACIENTE=" + str(list_pacientes[0][0]))

                            connection.commit()

                            st.write(":green[Usuario atualizado com sucesso!]")
            cancel = st.button("Voltar")
            if cancel:
                st.session_state['state_pacientes'] = 0
                st.experimental_rerun()

        if st.session_state['state_pacientes'] == 2:
            st.divider()
            st.title("Excluir paciente")
            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)
            if len(list_pacientes) ==0:
                st.error("Não foi possível encontrar pacientes")

            else:
                paciente = st.selectbox(
                    "Selecione o paciente ", list_pacientes)
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))
                motivo ="DESATIVADO"

                list_pacientes = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                Id_selected = c.fetchall()
                for i in Id_selected:
                    list_pacientes.append(i)

                enviar = st.button("Enviar")
                if enviar:
                    c.execute("UPDATE TABELA_PACIENTES_POSTO SET STATUS='" +
                            motivo + "' WHERE ID_PACIENTE=" + str(list_pacientes[0][0]))

                    connection.commit()

                    st.write(":green[Usuario atualizado com sucesso!]")
            cancel = st.button("Voltar")
            if cancel:
                st.session_state['state_pacientes'] = 0
                st.experimental_rerun()

    if(selected == "Gerenciar pedidos de coberturas"):
        logout = st.button("Logout")
        if(logout):
            LogOut()
        if st.session_state['state_pedidos'] == 0:
            st.divider()
            st.title("Gerenciar pedidos de coberturas")
            add_pedidos = st.button("Solicitar cobertura")
            if add_pedidos:
                st.session_state['state_pedidos'] = 1
                st.experimental_rerun()

            # verify_pedidos = st.button("Excluir pedidos")
            # if verify_pedidos:
            #     st.session_state['state_pedidos'] = 2
            #     st.experimental_rerun()

        if st.session_state['state_pedidos'] == 1:
            st.divider()
            st.title("Solicitar cobertura")
            
            c.execute(
                "SELECT MODELO_NOME FROM TABELA_COBERTURAS_POSTO;")

            list_tables = []
            tablesModelName = c.fetchall()

            for i in tablesModelName:
                value = i[0]
                list_tables.append(value)
            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(
        cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)

            c.execute(
                "SELECT NOME_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE ID_UNIDADE = " + str(st.session_state['ActualUser'])+" AND DATA_DE_ENCERRAMENTO is null;")
            Processos_ID = c.fetchall()
            list_processos = []
            for i in Processos_ID:
                value = i[0]
                list_processos.append(value)
            
            numero_de_pedidos = st.number_input(
                "Quantos pedidos deseja realizar ?", min_value=1, value=1, step=1, max_value=10)
            lista_quatidade = []
            lista_cobertura = []
            lista_processos = []
            list_ID = []
            lista_desc = []
            lista_pacientes = []
            if len(list_pacientes) ==0:
                st.error("Não foi possível encontrar pacientes")

            else:
                paciente = st.selectbox(
                "Selecione o paciente qual vai pedir os pedidos", list_pacientes)
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))
                col2, col3 = st.columns([3, 2])
                for i in range(0, int(numero_de_pedidos)):
                    with col2:
                        cobertura = st.selectbox(
                            "Selecione a cobertura que sera pedida no  "+str(i+1)+"º pedido ", list_tables)
                    with col3:
                        quantidade = st.number_input(
                            "Insira a quantidade da cobertura do "+str(i+1)+"º pedido", min_value=1, value=1, step=1)



                    lista_quatidade.append(quantidade)

                    c.execute(
                        "SELECT ID_ITEM FROM TABELA_COBERTURAS_POSTO WHERE MODELO_NOME = '" + cobertura + "';")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        lista_cobertura.append(i)

                    c.execute(
                        "SELECT DESCRICAO FROM TABELA_COBERTURAS_POSTO WHERE MODELO_NOME = '" + cobertura + "';")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        lista_desc.append(i)

                    c.execute(
                        "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '"+ str(paciente) + "';")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        lastName = i[0]
                        lista_pacientes.append(i)
                        
                    c.execute(
                        "SELECT ID_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE ID_PACIENTE = "+str(lastName)+" AND DATA_DE_ENCERRAMENTO IS NULL;")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        lista_processos.append(i)


                dateNow = datetime.now().strftime('%Y-%m-%d')
                enviar = st.button("Enviar")
                if enviar:
                    for i in range(0, len(lista_quatidade)):
                        c.execute("INSERT INTO TABELA_PEDIDOS_COBERTURAS_POSTO ( ID_MATERIAL, ID_DESC, QNT, PACIENTE_ID, ID_PROCESSO,DATA_ATUAL) VALUES (" + str(lista_cobertura[i][0]) + ",'" + str(lista_desc[i][0]) + "', " +
                                str(lista_quatidade[i]) + "," +
                                str(lista_pacientes[i][0])+"," +
                                str(lista_processos[i][0]) + ", STR_TO_DATE('" +
                                str(dateNow)+"','%Y-%m-%d'));")
                        connection.commit()
                    st.write(":green[Pedido cadastrado com sucesso!]")

            cancel = st.button("Voltar")
            if cancel:
                st.session_state['state_pedidos'] = 0
                st.experimental_rerun()

    if(selected == "Gerenciar processos de pacientes"):
        logout = st.button("Logout")
        if(logout):
            LogOut()
        if st.session_state['state_processos'] == 0:
            st.divider()
            st.title("Gerenciar processos de pacientes")
            verify_process = st.button("Editar processos de pacientes")
            if verify_process:
                st.session_state['state_processos'] = 3
                st.experimental_rerun()
            add_process = st.button("Adicionar processos de pacientes")
            if add_process:
                st.session_state['state_processos'] = 1
                st.experimental_rerun()
            remove_process = st.button("Fechar processos de pacientes")
            if remove_process:
                st.session_state['state_processos'] = 2
                st.experimental_rerun()

        if st.session_state['state_processos'] == 1:
            VerifyUser = False
            st.title("Adicionando processo de paciente")
            lista_pacientes = []

            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)
            if len(list_pacientes) ==0:
                st.error("Não foi possível encontrar pacientes")

            else:
                paciente = st.selectbox(
                    "Selecione o paciente ", list_pacientes)
                
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))

                
                c.execute("SELECT NOME_ETIOLOGIA FROM TABELA_ETIOLOGIAS_POSTO ;")
                etioList = c.fetchall()
                ListEtio = []
                for i in etioList:
                    ListEtio.append(i[0])

                c.execute("SELECT NOME_LOCAL_LESAO FROM TABELA_LOCAIS_LESAO ;")
                localList = c.fetchall()
                ListLocal = []
                for i in localList:
                    ListLocal.append(i[0])

                local = st.selectbox("Selecione o local da lesão",ListLocal)
                etio = st.selectbox("Selecione a etiologia da lesão",ListEtio)
                upload = st.file_uploader("Carregue o arquivo do processo")
                dateNow = datetime.now().strftime('%Y-%m-%d')
                if upload != None:
                    file_contais = upload.read()
                    pdf_base64 = base64.b64encode(file_contais).decode('utf-8')
                else: 
                    pdf_base64 = 0
                lista_pacientes = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                Id_selected = c.fetchall()
                for i in Id_selected:
                    lista_pacientes.append(i[0])

                listVerifyName = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PROCESSOS_POSTO WHERE DATA_DE_ENCERRAMENTO IS NULL ;")
                verifyed_selectedName = c.fetchall()
                for i in verifyed_selectedName:
                    listVerifyName.append(i[0])

                if lista_pacientes[0]  in listVerifyName:
                    st.error("Esse usuario já tem um processo em andamento")
                    VerifyUser = True
                
                elif((len(local) != 0) & (len(etio) != 0)):
                # st.write(pdf_base64)
                    enviar = st.button("Enviar")
                    if (enviar) & (VerifyUser == False)&(verificar_formato_localOuEtio(local,"Local da lesão")) & (verificar_formato_localOuEtio(etio,"Etiologia da lesão") & (pdf_base64 != 0)) :
                        c.execute("INSERT INTO TABELA_PROCESSOS_POSTO (LOCAL_DA_LESAO,ETIOLOGIA, PDF,ID_UNIDADE,DATA_DE_ABERTURA,DATA_DE_ENCERRAMENTO,NOME_PROCESSO,ID_PACIENTE) VALUES ('" + local + "','" + etio + "', '" +
                                pdf_base64 + "'," +
                                str(st.session_state['ActualUser']) + ", STR_TO_DATE('" +
                                str(dateNow) +
                                "','%Y-%m-%d'),"+"NULL"+",'"+local+"_"+etio+"_"+str(dateNow)+"_"+str(lista_pacientes[0])+
                                "','"+str(lista_pacientes[0])+"');"
                                )

                        connection.commit()
                        st.write(":green[Processo cadastrado com sucesso!]")
                    elif ((enviar) & (VerifyUser == False)&(verificar_formato_localOuEtio(local,"Local da lesão")) & (verificar_formato_localOuEtio(etio,"Etiologia da lesão")) & (pdf_base64 ==0)):
                        c.execute("INSERT INTO TABELA_PROCESSOS_POSTO (LOCAL_DA_LESAO,ETIOLOGIA,ID_UNIDADE,DATA_DE_ABERTURA,DATA_DE_ENCERRAMENTO,NOME_PROCESSO,ID_PACIENTE) VALUES ('" + local + "','" + etio + "', " +
                                str(st.session_state['ActualUser']) + ", STR_TO_DATE('" +
                                str(dateNow) +
                                "','%Y-%m-%d'),"+"NULL"+",'"+local+"_"+etio+"_"+str(dateNow)+str(lista_pacientes[0])+
                                "','"+str(lista_pacientes[0])+"');"
                        )
                        connection.commit()
                        st.write(":green[Processo cadastrado com sucesso!]")

            cancel = st.button("Voltar")

            if cancel:
                st.session_state['state_processos'] = 0
                st.experimental_rerun()

        if st.session_state['state_processos'] == 3:
            st.divider()
            st.title("Editando processo de paciente")
            lista_pacientes = []

            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)

            if len(list_pacientes) ==0 :

                st.error("Não foi possível encontrar nenhum processo") 

            else:
                c.execute(
                    "SELECT NOME_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE ID_UNIDADE = " + str(st.session_state['ActualUser'])+" AND DATA_DE_ENCERRAMENTO is null;")

                processo_ID = c.fetchall()
                lista_processo = []
                for i in processo_ID:
                    value = i[0]
                    lista_processo.append(value)


                process = st.selectbox(
                    "Selecione o processo ", lista_processo)

                paciente = st.selectbox(
                    "Selecione o paciente ", list_pacientes)
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))

                local = st.text_input("Insira o local da lesão")
                etio = st.text_input("Insira a etiologia da lesão")
                upload = st.file_uploader("Carregue o arquivo do processo")
                dateNow = datetime.now().strftime('%Y-%m-%d')
                if upload != None:
                    file_contais = upload.read()
                    pdf_base64 = base64.b64encode(file_contais).decode('utf-8')
                else: 
                    pdf_base64 = 0

                lista_pacientes = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                Id_selected = c.fetchall()
                for i in Id_selected:
                    lista_pacientes.append(i[0])

                lista_processo = []
                c.execute(
                    "SELECT ID_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE NOME_PROCESSO = '" + str(process) + "';")
                Id_selected = c.fetchall()
                for i in Id_selected:
                    lista_processo.append(i[0])

                listVerifyName = []
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PROCESSOS_POSTO WHERE DATA_DE_ENCERRAMENTO IS NULL ;")
                verifyed_selectedName = c.fetchall()
                for i in verifyed_selectedName:
                    listVerifyName.append(i[0])

                if((len(local) != 0) & (len(etio) != 0)):
                # st.write(pdf_base64)
                    enviar = st.button("Enviar")
                    if (enviar) &(verificar_formato_localOuEtio(local,"Local da lesão")) & (verificar_formato_localOuEtio(etio,"Etiologia da lesão")) &(len(local) != 0) & (len(etio) != 0) :
                        if pdf_base64 ==0:
                            c.execute("UPDATE TABELA_PROCESSOS_POSTO SET LOCAL_DA_LESAO='" + local +
                                    "', ETIOLOGIA='" + etio + 
                                    "', ID_UNIDADE=" + str(st.session_state['ActualUser']) +
                                    ", DATA_DE_ABERTURA=STR_TO_DATE('" + str(dateNow) + "', '%Y-%m-%d')" +
                                    ",NOME_PROCESSO = '"+local+"_"+etio+"_"+str(dateNow)+str(lista_pacientes[0]) +
                                    "',ID_PACIENTE ='"+str(lista_pacientes[0])+"'" + " WHERE ID_PROCESSO=" + str(lista_processo[0]) + ";")
                        elif  pdf_base64 !=0:   
                            c.execute("UPDATE TABELA_PROCESSOS_POSTO SET LOCAL_DA_LESAO='" + local +
                                    "', ETIOLOGIA='" + etio + "', PDF='" + pdf_base64 +
                                    "', ID_UNIDADE=" + str(st.session_state['ActualUser']) +
                                    ", DATA_DE_ABERTURA=STR_TO_DATE('" + str(dateNow) + "', '%Y-%m-%d')" +
                                    ",NOME_PROCESSO = '"+local+"_"+etio+"_"+str(dateNow)+str(lista_pacientes[0]) +
                                    "',ID_PACIENTE ='"+str(lista_pacientes[0])+"'" + " WHERE ID_PROCESSO=" + str(lista_processo[0]) + ";")

                        time.sleep(5)
                        connection.commit()

                        st.write(":green[Processo atualizado com sucesso!]")

            cancel = st.button("Voltar")

            if cancel:
                st.session_state['state_processos'] = 0
                st.experimental_rerun()

        if st.session_state['state_processos'] == 2:
            st.divider()
            st.title("Fechar processo de paciente")
            c.execute(
                "SELECT NOME FROM TABELA_PACIENTES_POSTO WHERE ID_UNIDADE = " + str(cookie_manager.get(cookie=cookiee)).split("|")[1]+" AND STATUS = 'ATIVO';")

            paciente_ID = c.fetchall()
            list_pacientes = []
            for i in paciente_ID:
                value = i[0]
                list_pacientes.append(value)
            if len(list_pacientes) ==0:
                st.error("Nenhum paciente encontrado")
            else:
                paciente = st.selectbox(
                "Selecione o paciente ", list_pacientes)
                c.execute("SELECT CNS FROM TABELA_PACIENTES_POSTO WHERE NOME = '" + str(paciente) + "';")
                cns = c.fetchall()
                st.subheader("CNS do paciente: "+str(cns[0][0]))
                c.execute(
                    "SELECT ID_PACIENTE FROM TABELA_PACIENTES_POSTO WHERE  NOME = '" + str(paciente) +"';")
                id_paciente = c.fetchall()
                c.execute(
                    "SELECT NOME_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE ID_UNIDADE = " + str(st.session_state['ActualUser'])+" AND DATA_DE_ENCERRAMENTO is null AND ID_PACIENTE = "+str(id_paciente[0][0])+";")
                processo_ID = c.fetchall()
                lista_processo = []
                for i in processo_ID:
                    value = i[0]
                    lista_processo.append(value)
                if len(lista_processo) == 0:
                    st.error("Esse paciente não tem processos")
                else:
                    process = st.selectbox(
                        "Selecione o processo ", lista_processo)
                    
                    c.execute(
                        "SELECT ID_PROCESSO FROM TABELA_PROCESSOS_POSTO WHERE NOME_PROCESSO = '" + str(process) + "';")
                    Id_selected = c.fetchall()
                    for i in Id_selected:
                        lista_processo.append(i[0])
                    motivo = st.selectbox(
                            "Selecione o motivo do fechamento do processo", ["ABANDONO", "ALTA/CURA", "INATIVO/MUDOUSE-SE", "OBITO", "OPTOU PELO PARTICULAR"])
                    dateNow = datetime.now().strftime('%Y-%m-%d')
                    enviar = st.button("Enviar")
                    if enviar:
                        c.execute("UPDATE TABELA_PROCESSOS_POSTO SET DATA_DE_ENCERRAMENTO =STR_TO_DATE('" + str(
                            dateNow) + "', '%Y-%m-%d')" + " AND MOTIVO_DO_ENCERRAMENTO = "+motivo+" WHERE ID_PROCESSO=" + str(lista_processo[0]) + ";")

                        time.sleep(5)
                        connection.commit()

                        st.write(":green[Processo encerrado com sucesso!]")

            cancel = st.button("Voltar")

            if cancel:
                st.session_state['state_processos'] = 0
                st.experimental_rerun()

elif st.session_state.Login2 == 3:
    st.error('Senha ou Usuario esta incorreto')
elif st.session_state.Login2 == 0:
    st.warning('Insira respectivamente o usuario e a senha como solicitado')
